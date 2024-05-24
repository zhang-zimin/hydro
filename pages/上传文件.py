import streamlit as st
import openpyxl
import base64


st.title('上传文件')
uploaded_file = st.file_uploader("上传文件", type=["csv", "json"])
if uploaded_file:
  st.write(f"你上传的文件是{uploaded_file.name}")



st.title('下载Excel')
xlsx_exporter = openpyxl.Workbook()
sheet = xlsx_exporter.active
sheet.cell(row=1, column=1).value = 'Thank'
sheet.cell(row=1, column=2).value = 'You'
sheet.cell(row=1, column=3).value = 'Python'
xlsx_exporter.save('results.xlsx')  # 注意！文件此时保存在内存中且为字节格式文件
data = open('results.xlsx', 'rb').read()  # 以只读模式读取且读取为二进制文件
b64 = base64.b64encode(data).decode('UTF-8')  # 解码并加密为base64
href = f'<a href="data:file/data;base64,{b64}" download="myresults.xlsx">Download xlsx file</a>'  # 定义下载链接，默认的下载文件名是myresults.xlsx
st.markdown(href, unsafe_allow_html=True)  # 输出到浏览器
